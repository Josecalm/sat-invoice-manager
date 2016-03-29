import openpyxl

from invoice import *

# The sheet's registries start at row 4
ROW_BEGIN = 4

# Sales Constants
DATE = 3
CLIENT_NAME = 4
SALES_SUB = 5
SALES_TAX = 6

# Purchase Constants
PROVIDER_RFC = 4
PROVIDER_NAME = 5
RATE_EXEMPT = 6
RATE_ZERO = 7
RATE_SIXTEEN = 8
PURCHASE_SUB = 9
PURCHASE_TAX = 10


def run(fileName, invoiceList, invoiceType):
	# Getting the xlsx document
	doc = openpyxl.load_workbook(fileName)
	listLength = len(invoiceList)
	# Working with cells
	if (invoiceType == "Ingresos"):
		test = doc.get_sheet_by_name('Test_Ing')
		r = ROW_BEGIN
		for invoice in invoiceList:
			c = test.cell(row = r, column = DATE)
			c.value = invoice.date
			c = test.cell(row = r, column = CLIENT_NAME)
			c.value = invoice.receiverName
			c = test.cell(row = r, column = SALES_SUB)
			c.value = invoice.subtotal
			c.number_format = '0.00'
			c = test.cell(row = r, column = SALES_TAX)
			c.value = invoice.tax
			c.number_format = '0.00'
			r += 1
	elif (invoiceType == "Egresos"):
		test = doc.get_sheet_by_name('Test_Egr')
		r = ROW_BEGIN
		for invoice in invoiceList:
			c = test.cell(row = r, column = DATE)
			c.value = invoice.date
			c = test.cell(row = r, column = PROVIDER_RFC)
			c.value = invoice.providerRfc
			c = test.cell(row = r, column = PROVIDER_NAME)
			c.value = invoice.providerName
			if (invoice.rate == 0):
				c = test.cell(row = r, column = RATE_ZERO)
				c.value = invoice.subtotal
				c.number_format = '0.00'
			else:
				c = test.cell(row = r, column = RATE_SIXTEEN)
				c.value = invoice.subtotal
				c.number_format = '0.00'
			c = test.cell(row = r, column = PURCHASE_SUB)
			c.value = invoice.subtotal
			c = test.cell(row = r, column = PURCHASE_TAX)
			c.value = invoice.tax
			r += 1
	else:
		print "Error al conseguir tipo de factura"
	doc.save('prueba.xlsx')
	exit = raw_input("Archivo generado. Presione enter para salir.")

